from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from typing import List, Optional
from datetime import datetime, timedelta, timezone
from database import get_db
import models, schemas
import httpx
import os
import json
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

router = APIRouter(prefix="/visitor", tags=["Visitors"])

def _date_filter(q, model, period: str):
    now = datetime.now(timezone.utc)
    if period == "day":
        start = now.replace(hour=0, minute=0, second=0, microsecond=0)
    elif period == "week":
        start = now - timedelta(days=now.weekday())
        start = start.replace(hour=0, minute=0, second=0, microsecond=0)
    elif period == "month":
        start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    elif period == "year":
        start = now.replace(month=1, day=1, hour=0, minute=0, second=0, microsecond=0)
    else:
        return q
    return q.filter(model.created_at >= start)

async def get_fcm_access_token() -> str:
    import google.auth.transport.requests
    from google.oauth2 import service_account
    sa_json = os.getenv("FIREBASE_SERVICE_ACCOUNT", "")
    if not sa_json:
        print("[FCM] No FIREBASE_SERVICE_ACCOUNT env var!")
        return ""
    try:
        sa_info     = json.loads(sa_json)
        credentials = service_account.Credentials.from_service_account_info(
            sa_info,
            scopes=["https://www.googleapis.com/auth/firebase.messaging"],
        )
        request = google.auth.transport.requests.Request()
        credentials.refresh(request)
        return credentials.token or ""
    except Exception as e:
        print(f"[FCM] Token error: {e}")
        return ""

async def send_fcm_notification(token: str, title: str, body: str, data: dict):
    if not token:
        print("[FCM] No token provided")
        return
    try:
        sa_json = os.getenv("FIREBASE_SERVICE_ACCOUNT", "")
        if not sa_json:
            print("[FCM] No FIREBASE_SERVICE_ACCOUNT env var!")
            return
        sa_info    = json.loads(sa_json)
        project_id = sa_info.get("project_id", "")
        print(f"[FCM] Project ID: {project_id}")
        url = f"https://fcm.googleapis.com/v1/projects/{project_id}/messages:send"

        access_token = await get_fcm_access_token()
        if not access_token:
            print("[FCM] Failed to get access token!")
            return

        print(f"[FCM] Got access token, sending notification...")
        payload = {
            "message": {
                "token"  : token,
                "android": {
                    "priority": "high",
                },
                "data": {
                    "title": title,
                    "body" : body,
                    **{k: str(v) for k, v in data.items()},
                },
            }
        }
        async with httpx.AsyncClient() as client:
            response = await client.post(
                url,
                json    = payload,
                headers = {
                    "Authorization": f"Bearer {access_token}",
                    "Content-Type" : "application/json",
                },
                timeout = 10,
            )
            print(f"[FCM] Response: {response.status_code} - {response.text}")
    except Exception as e:
        print(f"[FCM] Send failed: {e}")

@router.post("/create", response_model=schemas.VisitorResponse, status_code=201)
async def create_visitor(visitor: schemas.VisitorCreate, db: Session = Depends(get_db)):
    if visitor.logged_by is not None:
        guard = db.query(models.User).filter(
            models.User.id == visitor.logged_by).first()
        if not guard:
            raise HTTPException(status_code=404,
                detail=f"Guard with id {visitor.logged_by} not found")
        if guard.role not in ("security", "admin"):
            raise HTTPException(status_code=403,
                detail="Only security personnel can log visitors")

    new_visitor = models.Visitor(
        visitor_name    = visitor.visitor_name,
        phone           = visitor.phone,
        flat_no         = visitor.flat_no,
        visitor_type    = visitor.visitor_type,
        logged_by       = visitor.logged_by,
        checkin_time    = visitor.checkin_time,
        checkin_date    = visitor.checkin_date,
        is_prescheduled = visitor.is_prescheduled,
        society_id      = visitor.society_id,
        status          = "approved" if visitor.is_prescheduled else "pending",
    )
    db.add(new_visitor)
    db.commit()
    db.refresh(new_visitor)

    if not visitor.is_prescheduled:
        print(f"[FCM] Looking for resident at flat: {new_visitor.flat_no}")
        resident = db.query(models.User).filter(
            models.User.flat_no == new_visitor.flat_no,
            models.User.role    == "member",
            models.User.status  == "active",
        ).first()

        if resident:
            print(f"[FCM] Found resident: {resident.id}, token: {resident.fcm_token}")
        else:
            print(f"[FCM] No resident found for flat: {new_visitor.flat_no}")

        if resident and resident.fcm_token:
            print(f"[FCM] Sending notification...")
            await send_fcm_notification(
                token = resident.fcm_token,
                title = "Visitor at Gate 🔔",
                body  = f"{new_visitor.visitor_name} ({new_visitor.visitor_type}) is at the gate. Approve?",
                data  = {
                    "visitor_id": str(new_visitor.id),
                    "flat_no"   : new_visitor.flat_no,
                    "type"      : "visitor_request",
                },
            )
        elif resident and not resident.fcm_token:
            print(f"[FCM] Resident found but no FCM token!")

    return new_visitor

@router.post("/preschedule", response_model=schemas.VisitorResponse, status_code=201)
def preschedule_visitor(visitor: schemas.VisitorCreate, db: Session = Depends(get_db)):
    new_visitor = models.Visitor(
        visitor_name    = visitor.visitor_name,
        phone           = visitor.phone,
        flat_no         = visitor.flat_no,
        visitor_type    = visitor.visitor_type,
        logged_by       = visitor.logged_by,
        checkin_time    = visitor.checkin_time,
        checkin_date    = visitor.checkin_date,
        is_prescheduled = True,
        society_id      = visitor.society_id,
        status          = "approved",
    )
    db.add(new_visitor)
    db.commit()
    db.refresh(new_visitor)
    return new_visitor

@router.post("/checkout", response_model=schemas.VisitorResponse)
def checkout_visitor(data: schemas.VisitorCheckout, db: Session = Depends(get_db)):
    visitor = db.query(models.Visitor).filter(
        models.Visitor.id == data.visitor_id).first()
    if not visitor:
        raise HTTPException(status_code=404, detail="Visitor not found")
    if visitor.checkout_time:
        raise HTTPException(status_code=409, detail="Visitor already checked out")
    visitor.checkout_time = data.checkout_time
    visitor.checkout_date = data.checkout_date
    db.commit()
    db.refresh(visitor)
    return visitor

@router.post("/approve", response_model=schemas.VisitorResponse)
def approve_visitor(data: schemas.VisitorApprove, db: Session = Depends(get_db)):
    visitor = db.query(models.Visitor).filter(
        models.Visitor.id == data.visitor_id).first()
    if not visitor:
        raise HTTPException(status_code=404, detail="Visitor not found")
    if visitor.status != "pending":
        raise HTTPException(status_code=409,
            detail=f"Visitor request is already '{visitor.status}'")
    visitor.status = data.action
    db.commit()
    db.refresh(visitor)
    return visitor

@router.get("/export/excel")
def export_visitors_excel(
    from_date  : Optional[str] = Query(None),
    to_date    : Optional[str] = Query(None),
    society_id : Optional[int] = Query(None),
    flat_no    : Optional[str] = Query(None),
    status     : Optional[str] = Query(None),
    db         : Session       = Depends(get_db),
):
    q = db.query(models.Visitor)

    if from_date:
        try:
            fd = datetime.strptime(from_date, "%Y-%m-%d").replace(tzinfo=timezone.utc)
            q  = q.filter(models.Visitor.created_at >= fd)
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid from_date. Use YYYY-MM-DD")

    if to_date:
        try:
            td = datetime.strptime(to_date, "%Y-%m-%d").replace(
                hour=23, minute=59, second=59, tzinfo=timezone.utc)
            q  = q.filter(models.Visitor.created_at <= td)
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid to_date. Use YYYY-MM-DD")

    if society_id: q = q.filter(models.Visitor.society_id == society_id)
    if flat_no:    q = q.filter(models.Visitor.flat_no     == flat_no)
    if status:     q = q.filter(models.Visitor.status      == status)

    visitors = q.order_by(models.Visitor.created_at.desc()).all()

    # ── Build Excel workbook ───────────────────────────────────────────────────
    wb = Workbook()
    ws = wb.active
    ws.title = "Visitor Report"

    header_fill  = PatternFill("solid", fgColor="1A6B3A")
    header_font  = Font(bold=True, color="FFFFFF", size=11)
    alt_fill     = PatternFill("solid", fgColor="E8F5EC")
    center_align = Alignment(horizontal="center", vertical="center")
    left_align   = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin_border  = Border(
        left  =Side(style="thin", color="CCCCCC"),
        right =Side(style="thin", color="CCCCCC"),
        top   =Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )

    # Row 1 — Title
    ws.merge_cells("A1:L1")
    ws["A1"]           = "VMF — Visitor Management Report"
    ws["A1"].font      = Font(bold=True, size=14, color="1A6B3A")
    ws["A1"].alignment = center_align
    ws.row_dimensions[1].height = 30

    # Row 2 — Date range info
    ws.merge_cells("A2:L2")
    ws["A2"]           = f"Period: {from_date or 'All'} to {to_date or 'All'}  |  Total Records: {len(visitors)}"
    ws["A2"].font      = Font(size=10, color="666666", italic=True)
    ws["A2"].alignment = center_align
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 5

    # Row 4 — Headers
    headers    = ["Sr No","Visitor Name","Phone","Flat No","Visitor Type",
                  "Status","Check-in Date","Check-in Time",
                  "Check-out Date","Check-out Time","Pre-Scheduled","Logged At"]
    col_widths = [7, 22, 14, 12, 14, 12, 14, 12, 14, 12, 13, 20]

    for col, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell           = ws.cell(row=4, column=col, value=header)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = center_align
        cell.border    = thin_border
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.row_dimensions[4].height = 22

    status_colors = {"approved":"2ECC71","rejected":"E74C3C","pending":"F39C12"}

    for i, v in enumerate(visitors, 1):
        row   = 4 + i
        fill  = alt_fill if i % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
        created = v.created_at.strftime("%d-%m-%Y %H:%M") if v.created_at else ""

        data_row = [
            i,
            v.visitor_name,
            v.phone,
            v.flat_no,
            v.visitor_type,
            v.status.upper(),
            v.checkin_date  or "—",
            v.checkin_time  or "—",
            v.checkout_date or "—",
            v.checkout_time or "—",
            "Yes" if v.is_prescheduled else "No",
            created,
        ]
        for col, value in enumerate(data_row, 1):
            cell           = ws.cell(row=row, column=col, value=value)
            cell.fill      = fill
            cell.border    = thin_border
            cell.alignment = center_align if col in [1,6,7,8,9,10,11] else left_align
            if col == 6:
                cell.font = Font(bold=True, color=status_colors.get(v.status, "000000"))
        ws.row_dimensions[row].height = 18

    ws.freeze_panes = "A5"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"VMF_Visitors_{from_date or 'All'}_{to_date or 'All'}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@router.get("/list", response_model=List[schemas.VisitorResponse])
def list_visitors(
    flat_no   : Optional[str] = Query(None),
    status    : Optional[str] = Query(None),
    period    : Optional[str] = Query(None),
    society_id: Optional[int] = Query(None),
    skip      : int           = Query(0, ge=0),
    limit     : int           = Query(20, ge=1, le=100),
    db        : Session       = Depends(get_db),
):
    q = db.query(models.Visitor)
    if flat_no:    q = q.filter(models.Visitor.flat_no    == flat_no)
    if status:     q = q.filter(models.Visitor.status     == status)
    if period:     q = _date_filter(q, models.Visitor, period)
    if society_id: q = q.filter(models.Visitor.society_id == society_id)
    return q.order_by(
        models.Visitor.created_at.desc()).offset(skip).limit(limit).all()

@router.get("/dashboard/metrics")
def dashboard_metrics(
    period    : str           = Query("day"),
    flat_no   : Optional[str] = Query(None),
    society_id: Optional[int] = Query(None),
    db        : Session       = Depends(get_db),
):
    q = db.query(models.Visitor)
    if flat_no:    q = q.filter(models.Visitor.flat_no    == flat_no)
    if society_id: q = q.filter(models.Visitor.society_id == society_id)
    q = _date_filter(q, models.Visitor, period)
    all_visitors = q.all()
    return {
        "period"      : period,
        "total"       : len(all_visitors),
        "pending"     : sum(1 for v in all_visitors if v.status == "pending"),
        "approved"    : sum(1 for v in all_visitors if v.status == "approved"),
        "rejected"    : sum(1 for v in all_visitors if v.status == "rejected"),
        "prescheduled": sum(1 for v in all_visitors if v.is_prescheduled),
        "checked_out" : sum(1 for v in all_visitors if v.checkout_time),
    }

@router.get("/{visitor_id}", response_model=schemas.VisitorResponse)
def get_visitor(visitor_id: int, db: Session = Depends(get_db)):
    visitor = db.query(models.Visitor).filter(
        models.Visitor.id == visitor_id).first()
    if not visitor:
        raise HTTPException(status_code=404, detail="Visitor not found")
    return visitor